"""
Excel export — generates a professional multi-sheet workbook.

Sheets:
1. Project_Info   — metadata and settings
2. Members        — member strip definitions
3. Loads          — raw load input
4. Overlap_Calculations — full results with d1/d2
5. STAAD_Export   — formatted STAAD commands
6. Summary        — quick QA statistics
"""

from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import (
    ProjectSettings, MemberSegment, LoadPatch, OverlapResult, OverlapSummary
)
from app.core.staad_export import generate_staad_text


# ─── Styles ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")

TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def generate_excel(
    settings: ProjectSettings,
    members: list[MemberSegment],
    loads: list[LoadPatch],
    overlaps: list[OverlapResult],
    summary: OverlapSummary,
    precision: int = 2,
) -> io.BytesIO:
    """
    Generate a professional Excel workbook with 6 sheets.

    Returns:
        BytesIO buffer containing the .xlsx file
    """
    wb = Workbook()

    # Sheet 1: Project Info
    _write_project_info(wb, settings)

    # Sheet 2: Members
    _write_members(wb, members, precision)

    # Sheet 3: Loads
    _write_loads(wb, loads, precision)

    # Sheet 4: Overlap Calculations
    _write_overlaps(wb, overlaps, precision)

    # Sheet 5: STAAD Export
    _write_staad(wb, overlaps, precision)

    # Sheet 6: Summary
    _write_summary(wb, summary)

    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_project_info(wb: Workbook, settings: ProjectSettings):
    ws = wb.create_sheet("Project_Info")

    ws["A1"] = "BRIDGE LOAD SEGMENTATION REPORT"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    info = [
        ("Project Name", settings.project_name),
        ("Bridge Name", settings.bridge_name),
        ("Engineer", settings.engineer),
        ("Date", settings.project_date),
        ("Structure Type", settings.structure_type.value),
        ("Total Width", f"{settings.total_width} {settings.units.value}"),
        ("Reference Axis", settings.reference_axis.value),
        ("Decimal Precision", str(settings.decimal_precision)),
        ("Units", settings.units.value),
        ("Fill Depth", f"{settings.fill_depth} {settings.units.value}"),
        ("Slab Thickness", f"{settings.slab_thickness} {settings.units.value}"),
        ("Wall Thickness", f"{settings.wall_thickness} {settings.units.value}"),
        ("Overhang Policy", settings.overhang_policy.value),
        ("Comments", settings.comments),
    ]

    for i, (label, value) in enumerate(info, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = SUBTITLE_FONT
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = DATA_FONT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40


def _write_members(wb: Workbook, members: list[MemberSegment], precision: int):
    ws = wb.create_sheet("Members")

    headers = ["Member ID", "Start", "End", "Width", "Group", "Label"]
    _write_header_row(ws, headers)

    for i, m in enumerate(members, start=2):
        row = [m.id, round(m.start, precision), round(m.end, precision),
               round(m.width, precision), m.group.value, m.label]
        _write_data_row(ws, i, row)

    _autofit_columns(ws, headers)
    ws.freeze_panes = "A2"


def _write_loads(wb: Workbook, loads: list[LoadPatch], precision: int):
    ws = wb.create_sheet("Loads")

    headers = ["Load ID", "Load Case", "Type", "Start", "End",
               "Intensity", "Intensity End", "Direction", "Notes"]
    _write_header_row(ws, headers)

    for i, lo in enumerate(loads, start=2):
        row = [
            lo.id, lo.load_case, lo.load_type.value,
            round(lo.start, precision), round(lo.end, precision),
            round(lo.intensity, precision),
            round(lo.intensity_end, precision) if lo.intensity_end is not None else "",
            lo.direction.value, lo.notes,
        ]
        _write_data_row(ws, i, row)

    _autofit_columns(ws, headers)
    ws.freeze_panes = "A2"


def _write_overlaps(wb: Workbook, overlaps: list[OverlapResult], precision: int):
    ws = wb.create_sheet("Overlap_Calculations")

    headers = [
        "Load ID", "Load Case", "Member ID",
        "Member Start", "Member End",
        "Load Start (Global)", "Load End (Global)",
        "Overlap Start (Global)", "Overlap End (Global)",
        "Front Distance (d1)", "Back Distance (d2)",
        "Loaded Length", "Intensity", "Intensity End",
        "Direction", "STAAD Format", "Notes",
    ]
    _write_header_row(ws, headers)

    for i, r in enumerate(overlaps, start=2):
        row = [
            r.load_id, r.load_case, r.member_id,
            r.member_start, r.member_end,
            r.load_start_global, r.load_end_global,
            r.overlap_start_global, r.overlap_end_global,
            r.front_distance, r.back_distance,
            r.loaded_length, r.intensity,
            r.intensity_end if r.intensity_end is not None else "",
            r.direction.value, r.staad_format.value, r.notes,
        ]
        _write_data_row(ws, i, row)

    _autofit_columns(ws, headers)
    ws.freeze_panes = "A2"


def _write_staad(wb: Workbook, overlaps: list[OverlapResult], precision: int):
    ws = wb.create_sheet("STAAD_Export")

    ws["A1"] = "STAAD.Pro Member Load Commands"
    ws["A1"].font = TITLE_FONT

    staad_text = generate_staad_text(overlaps, precision)
    for i, line in enumerate(staad_text.split("\n"), start=3):
        ws[f"A{i}"] = line
        ws[f"A{i}"].font = Font(name="Consolas", size=10)

    ws.column_dimensions["A"].width = 80


def _write_summary(wb: Workbook, summary: OverlapSummary):
    ws = wb.create_sheet("Summary")

    ws["A1"] = "QA SUMMARY"
    ws["A1"].font = TITLE_FONT

    stats = [
        ("Total Members", summary.total_members),
        ("Total Loads", summary.total_loads),
        ("Affected Members", summary.affected_members),
        ("Total Overlap Rows", summary.total_overlap_rows),
    ]

    for i, (label, value) in enumerate(stats, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = SUBTITLE_FONT
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = DATA_FONT

    # Per-load stats
    row = len(stats) + 5
    ws[f"A{row}"] = "Loaded Width per Load"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    ws[f"A{row}"] = "Load ID"
    ws[f"B{row}"] = "Total Width"
    ws[f"C{row}"] = "Affected Members"
    for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for load_id, width in summary.total_loaded_width_by_load.items():
        row += 1
        ws[f"A{row}"] = load_id
        ws[f"B{row}"] = width
        ws[f"C{row}"] = summary.affected_members_by_load.get(load_id, 0)

    # Per-case stats
    row += 2
    ws[f"A{row}"] = "Loaded Width per Load Case"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    ws[f"A{row}"] = "Case ID"
    ws[f"B{row}"] = "Total Width"
    for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for case_id, width in summary.total_loaded_width_by_case.items():
        row += 1
        ws[f"A{row}"] = case_id
        ws[f"B{row}"] = width

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _write_header_row(ws, headers: list[str]):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _write_data_row(ws, row: int, values: list):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = THIN_BORDER
        if row % 2 == 0:
            cell.fill = ALT_ROW_FILL


def _autofit_columns(ws, headers: list[str]):
    for i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        # Set width based on header length with some padding
        width = max(len(str(header)) + 4, 12)
        ws.column_dimensions[col_letter].width = width
